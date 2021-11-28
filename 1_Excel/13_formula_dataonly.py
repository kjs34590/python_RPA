from openpyxl import load_workbook
wb = load_workbook("./PYTHON_WORKS/python_RPA/1_Excel/sample_formula.xlsx")
ws = wb.active

# 수식 그대로 가져오기
for row in ws.values:
    for cell in row:
        print(cell)

# 계산된 값 가져오기
wb = load_workbook("./PYTHON_WORKS/python_RPA/1_Excel/sample_formula.xlsx", data_only=True)
ws = wb.active
# openpyxl에서는 수식이 evaluate되지 않았기 때문에 None 반환
# 엑셀 파일을 직접 열어서 저장한 후 실행하면 적용됨
for row in ws.values:
    for cell in row:
        print(cell)
