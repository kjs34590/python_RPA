from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 한 줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"]) # list or tuple 형태로 입력
for i in range(1, 11): # 10명의 점수 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # B컬럼(영어)의 데이터를 모두 가져와 col_B에 저장
print(col_B)

for cell in col_B:
    print(cell.value)

col_range = ws["B:C"] # B부터 C까지의 컬럼 데이터 가져오기
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1] # 첫 번째 row만 가져오기
for cell in row_title:
    print(cell.value)

row_range = ws[1:6]
for rows in row_range:
    for cell in rows:
        print(cell.value)
    print()


from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ") # 값 가져오기
        # print(cell.coordinate, end=" ") # 각 셀 정보 가져오기
        xy = coordinate_from_string(cell.coordinate)
        # print(xy, end=" ")
        print(xy[0], end="") # A
        print(xy[1], end=" ") # 1
    print()


# 전체 rows
print(tuple(ws.rows))
for row in tuple(ws.rows):
    print(row[1].value)

# 전체 columns
print(tuple(ws.columns))
for col in tuple(ws.columns):
    print(col[1].value)

for row in ws.iter_rows(): # 전체 row
    print(row[1].value)

for col in ws.iter_cols():
    print(col[1].value)

# 슬라이싱이 가능하다.
for row in ws.iter_rows(min_row=1, max_row=5): # 1~5 row
    print(row[1].value)

wb.save("./PYTHON_WORKS/python_RPA/1_Excel/cell_range.xlsx") 