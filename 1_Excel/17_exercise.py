# Quiz ) 당신은 모대학의 학과 조교입니다.
# 이번 학기 성적을 입력해야 하는 과목의 점수 비중은 다음과 같습니다.

# 출석 10
# 퀴즈1 10
# 퀴즈2 10
# 중간고사 20
# 기말고사 30
# 프로젝트 20
# ----------- 총합 100

# [ 현재까지 작성된 최종 성적 데이터 ]
# 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
# 1, 10, 8, 5, 14, 26, 12
# 2, 7, 3, 7, 15, 24, 18
# 3, 9, 5, 8, 8, 12, 4
# 4, 7, 8, 7, 17, 21, 18
# 5, 7, 8, 7, 16, 25, 15
# 6, 3, 5, 8, 8, 17, 0
# 7, 4, 9, 10, 16, 27, 18
# 8, 6, 6, 6, 15, 19, 17
# 9, 10, 10, 9, 19, 30, 19
# 10, 9, 8, 8, 20, 25, 20

# 그런데 학생들의 최종 성적을 검토하는 과정에서, 퀴즈2 문제에 오류가 있음을 발견해
# 모두 만점 처리를 해야 합니다. 작성한 최종 성적 데이터를 기준으로 아래와 같이 수정하세요.

# 1. 퀴즈2 점수를 10으로 수정
# 2. H열에 총점(SUM 이용), I열에 성적 정보 추가
# - 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지는 D
# 3. 출석이 5 미만인 학생은 총점에 관계 없이 F

# 최종 산출물 파일명 : scores.xlsx


from openpyxl import Workbook
wb = Workbook()
ws = wb.active # 현재 활성화된 sheet 가져오기
ws.title = "Scores" # sheet 이름 변경

ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트", "총점", "성적"])

scores = [
(1, 10, 8, 5, 14, 26, 12),
(2, 7, 3, 7, 15, 24, 18),
(3, 9, 5, 8, 8, 12, 4),
(4, 7, 8, 7, 17, 21, 18),
(5, 7, 8, 7, 16, 25, 15),
(6, 3, 5, 8, 8, 17, 0),
(7, 4, 9, 10, 16, 27, 18),
(8, 6, 6, 6, 15, 19, 17),
(9, 10, 10, 9, 19, 30, 19),
(10, 9, 8, 8, 20, 25, 20)
]

for s in scores :
    ws.append(s)

# 퀴즈 2 점수 10점으로 일괄 변경
for idx, cell in enumerate(ws["D"]): # enumerate는 첫 번째 줄 제외를 위해
    if idx == 0:
        continue
    cell.value = 10

for i, s in enumerate(scores, start=2):
    print(i, s)

# H열에 총점(SUM 이용) 추가
for idx, score in enumerate(scores, start=2): # 2번째부터 엑셀 데이터가 시작하므로
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)
    # 총점 90 이상 A, 80 이상 B, 70 이상 C, 나머지 D로 성적(I열) 추가
    grade = None
    sum_val = sum(score[1:]) - score[3] + 10 # evaluate 되지 않았으므로, 계산을 위한 변수 생성
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"
    # 출석이 5 미만이면 무조건 F
    if score[1] < 5:
        grade = "F"
    
    ws.cell(row=idx, column=9).value = grade

wb.save("./PYTHON_WORKS/python_RPA/1_Excel/scores.xlsx")