from openpyxl import load_workbook
wb = load_workbook("./PYTHON_WORKS/python_RPA/1_Excel/sample_merge.xlsx")
ws = wb.active

# B2:D2 병합돼 있던 셀을 병합 해제
ws.unmerge_cells("B2:D2")
wb.save("./PYTHON_WORKS/python_RPA/1_Excel/sample_unmerge.xlsx")