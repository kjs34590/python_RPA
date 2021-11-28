from openpyxl import load_workbook
wb = load_workbook("./PYTHON_WORKS/python_RPA/1_Excel/cell_range.xlsx")
ws = wb.active

# # 번호 영어 수학 -> 번호 수학 영어
# ws.move_range("C1:C11", rows=5, cols=-1) # 컬럼은 왼쪽 1칸, 로우는 5칸 아래로 이동 (덮어쓰기)

# 번호 영어 수학 -> 번호 국어 영어 수학
ws.move_range("B1:C11", rows=0, cols=1) # 대상을 col 1만큼 이동
ws["B1"].value = "국어" # B1 셀에 '국어' 입력

wb.save("./PYTHON_WORKS/python_RPA/1_Excel/add_korean.xlsx")