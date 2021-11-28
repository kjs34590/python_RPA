from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 시트 생성 (기본 이름으로)
ws.title = "My_Sheet" # 시트 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # 색상값 입력 시 탭 색상 변경

ws1 = wb.create_sheet("Your_Sheet") # 생성하면서 타이틀을 바로 줄 수 있다.
ws2 = wb.create_sheet("New_Sheet", 2) # 2번째 인덱스에 시트 생성

new_ws = wb["New_Sheet"] # Dict 형태로 시트에 접근

print(wb.sheetnames)

# 시트 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied_Sheet"

wb.save("./PYTHON_WORKS/python_RPA/1_Excel/sample.xlsx")
