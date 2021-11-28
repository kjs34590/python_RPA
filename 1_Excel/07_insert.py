from openpyxl import load_workbook
wb = load_workbook("./PYTHON_WORKS/python_RPA/1_Excel/cell_range.xlsx")
ws = wb.active
# ws.insert_rows(8) # 8번째 줄 비우기(한 줄 삽입)
ws.insert_rows(8, 5) # 8번째 줄부터 5줄 삽입
wb.save("./PYTHON_WORKS/python_RPA/1_Excel/insert_rows.xlsx") # 새 파일에 저장


wb = load_workbook("./PYTHON_WORKS/python_RPA/1_Excel/cell_range.xlsx")
ws = wb.active
ws.insert_cols(2,3) # B열에 3개의 컬럼 추가
wb.save("./PYTHON_WORKS/python_RPA/1_Excel/insert_cols.xlsx") # 새 파일에 저장
