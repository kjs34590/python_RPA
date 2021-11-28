from openpyxl import Workbook
import datetime

wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()
ws["A2"] = "=SUM(1, 2, 3)" # 6 (합계)
ws["A3"] = "=AVERAGE(1, 2, 3)" # 2 (평균)

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)" # 30

wb.save("./PYTHON_WORKS/python_RPA/1_Excel/sample_formula.xlsx")