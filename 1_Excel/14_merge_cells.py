from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 셀 병합하기
ws.merge_cells("B2:D2") # B2~D2 셀을 합친다.
ws["B2"].value = "Merged Cell"

wb.save("./PYTHON_WORKS/python_RPA/1_Excel/sample_merge.xlsx")