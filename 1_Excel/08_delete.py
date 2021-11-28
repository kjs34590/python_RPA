from openpyxl import load_workbook
wb = load_workbook("./PYTHON_WORKS/python_RPA/1_Excel/cell_range.xlsx")
ws = wb.active

# ws.delete_rows(8) # 7번 학생 데이터 삭제
ws.delete_rows(8, 3) # 7, 8, 9번 학생 데이터 삭제
wb.save("./PYTHON_WORKS/python_RPA/1_Excel/delete_rows.xlsx")


wb = load_workbook("./PYTHON_WORKS/python_RPA/1_Excel/cell_range.xlsx")
ws = wb.active
# ws.delete_cols(2) # 영어 성적 컬럼 삭제
ws.delete_cols(2, 2) # 영어, 수학 성적 컬럼 삭제
wb.save("./PYTHON_WORKS/python_RPA/1_Excel/delete_cols.xlsx")

