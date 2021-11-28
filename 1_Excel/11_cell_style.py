from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = load_workbook("./PYTHON_WORKS/python_RPA/1_Excel/cell_range.xlsx")
ws = wb.active

a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

# A열 너비 5로 설정
ws.column_dimensions["A"].width = 5

# 1행의 높이를 50으로 설정
ws.row_dimensions[1].height = 25

# 폰트 변경
a1.font = Font(color="000099", italic=True, bold=True)
b1.font = Font(color="4C0099", name="Arial", strike=True)
c1.font = Font(color="99004C", size=16, underline="single")

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀을 초록색으로 변경
for row in ws.rows:
    for cell in row:
        # 가운데 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column == 1: # A (번호) 열은 제외
            continue
        # cell이 정수형 데이터이고 90점보다 높다는 조건 설정
        if isinstance(cell.value, int) and cell.value > 90 :
            cell.fill = PatternFill(fgColor="006600", fill_type="solid")
            cell.font = Font(color="FFFFFF") # 폰트 색 변경

# 틀 고정
ws.freeze_panes = "B2" # B2 기준으로 틀 고정 (A열과 1행이 고정됨)

wb.save("./PYTHON_WORKS/python_RPA/1_Excel/cell_style.xlsx")