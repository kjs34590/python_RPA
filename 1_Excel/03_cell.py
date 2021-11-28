from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "MySheet"

# 셀에 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])
print(ws["A1"].value)
print(ws["A10"].value) # 값이 없을 땐 None 출력

ws.cell(row=1, column=1) # ws["A1"]과 동일

c = ws.cell(column=3, row=1, value=10) # ws["C1"] = 10 과 동일
print(c.value) # ws["C1"].value를 print한 것과 동일

from random import *

# 반복문을 이용해서 랜덤 숫자 채우기
index = 1
for x in range(1, 11): # 10개 row
    for y in range(1, 11): # 10개 column
        # ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 사이 숫자 랜덤 입력
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("./PYTHON_WORKS/python_RPA/1_Excel/sample.xlsx")
